from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup
import lxml
import re

url = 'https://nos.nl/'
r = requests.get(url)
rtext = r.text

soup = BeautifulSoup(rtext, features="lxml")
titles = soup.find_all("h2")

nos = Workbook()
activeSheet = nos.active

column = "a"

row = 1
count = 0
cell = ""
for x in titles:
    o = x.text
    m = o.split()
    for i in m:
        regex = re.compile('[^a-zA-Z]')
        i = regex.sub('', i)
        if len(i) > 3:
            cell = column + str(row)
            activeSheet[cell] = i
            row += 1
column2 = "b"
row = 1
rowWord = 1
count = 0
cell = ""
go = True
while go:
    wordCheck = True
    wordCount = 0
    pasteCell = column2 + str(row)
    checkCell = column + str(row)
    checkCellValue = activeSheet[checkCell].value
    
    iterate = 1
    while wordCheck:
        cell = column + str(iterate)
        if activeSheet[cell].value == None:
            wordCheck = False
        if activeSheet[checkCell].value == activeSheet[cell].value:
            wordCount += 1
            print(wordCount)
        iterate += 1
        
    activeSheet[pasteCell] = wordCount
    row += 1
    row2 = row + 1
    checkCell = column + str(row2)

    if activeSheet[checkCell].value == None:
        go = False



    
nos.save("nos.xlsx")

#wb = Workbook()

#ws = wb.active

#ws['a1'] = 9
#ws['b1'] = 3
#ws['c1'] = ws['a1'].value * ws['b1'].value

#wb.save("Test.xlsx")


